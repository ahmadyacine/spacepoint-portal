"""
payment_service.py
------------------
Handles:
  - Generating the Facilitator Payment Letter as PDF (ReportLab)
  - Embedding the instructor's canvas digital signature into the PDF
  - Parsing the bulk Excel import (.xlsx, 2-sheet format)
"""

import os
import io
import base64
from datetime import datetime
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
from reportlab.platypus.flowables import Image as RLImage
from reportlab.pdfgen import canvas as rl_canvas

from sqlalchemy.orm import Session
from app.models.payment import (
    PaymentLetter, PaymentLetterStatus, InstructorBankDetails, PortalSetting
)
from app.models.user import User

# ── paths ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_STATIC = os.path.join(_HERE, "..", "static")
_LOGO_PATH = os.path.join(_STATIC, "SpacePoint logo.png")
_UPLOADS_DIR = os.path.join(_HERE, "..", "uploads", "payment_letters")
os.makedirs(_UPLOADS_DIR, exist_ok=True)

# Brand colours
BRAND_DARK = colors.HexColor("#1a1135")
BRAND_PURPLE = colors.HexColor("#653f84")
BRAND_LIGHT_PURPLE = colors.HexColor("#f3eeff")
HEADER_BG = colors.HexColor("#231134")
TABLE_HEADER_GREY = colors.HexColor("#d8d8d8")
WHITE = colors.white
BLACK = colors.black


# ── helper: get portal setting ─────────────────────────────────────────────
def _get_setting(db: Session, key: str, default: str = "") -> str:
    row = db.query(PortalSetting).filter(PortalSetting.key == key).first()
    return row.value if row and row.value else default


# ── helper: draw full bleed header ─────────────────────────────────────────
def draw_full_header(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#231134"))
    # A4: 210 x 297 mm. Rect height is 18mm, so y ranges from 279 to 297.
    canvas.rect(0, 279 * mm, 210 * mm, 18 * mm, fill=1, stroke=0)
    
    if os.path.exists(_LOGO_PATH):
        try:
            canvas.drawImage(_LOGO_PATH, 20 * mm, 282 * mm, width=51.6 * mm, height=12 * mm, mask='auto')
        except Exception:
            # Fallback text
            canvas.setFillColor(colors.white)
            canvas.setFont("Helvetica-Bold", 16)
            canvas.drawString(20 * mm, 284 * mm, "SPACE.")
    else:
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawString(20 * mm, 284 * mm, "SPACE.")
        
    canvas.restoreState()


# ── main PDF builder ───────────────────────────────────────────────────────
def generate_payment_letter_pdf(letter_id: int, db: Session) -> Optional[str]:
    """
    Generates a Facilitator Payment Letter PDF for the given payment_letter.
    Saves the file to uploads/payment_letters/ and updates letter.pdf_path.
    Returns the absolute path to the generated PDF, or None on failure.
    """
    letter = db.query(PaymentLetter).filter(PaymentLetter.id == letter_id).first()
    if not letter:
        return None

    instructor = db.query(User).filter(User.id == letter.instructor_user_id).first()
    if not instructor:
        return None

    bank = db.query(InstructorBankDetails).filter(
        InstructorBankDetails.user_id == letter.instructor_user_id
    ).first()

    # Portal settings — admin signatory
    admin_name = _get_setting(db, "admin_signatory_name", "SpacePoint Admin")
    admin_sig_path = _get_setting(db, "admin_signature_path", "")

    # Compute totals
    sessions = letter.sessions
    addons = letter.addons
    base_total = sum(s.compensation_aed for s in sessions)
    addons_total = sum(a.amount_aed for a in addons)
    grand_total = base_total + addons_total

    # File path
    safe_name = "".join(c for c in instructor.name if c.isalnum() or c in " _-").strip().replace(" ", "_")
    filename = f"PaymentLetter_{safe_name}_{letter.id}.pdf"
    output_path = os.path.join(_UPLOADS_DIR, filename)

    # Letter date
    letter_date = letter.letter_date or datetime.now().strftime("%d/%m/%Y")

    # ── Build PDF ────────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=10 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # ── Header bar spacer ────────────────────────────────────────────────
    story.append(Spacer(1, 18 * mm))

    # ── Title ────────────────────────────────────────────────────────────
    title_style = ParagraphStyle("title", fontSize=13, fontName="Helvetica-Bold",
                                 textColor=BRAND_DARK, alignment=TA_CENTER,
                                 underline=True, leading=18)
    story.append(Paragraph("<u><b>SpacePoint's Facilitator Payment Letter</b></u>", title_style))
    story.append(Spacer(1, 5 * mm))

    # ── Meta block ───────────────────────────────────────────────────────
    normal = ParagraphStyle("normal", fontSize=10, fontName="Helvetica", leading=16)
    bold_label = ParagraphStyle("bold_label", fontSize=10, fontName="Helvetica-Bold", leading=16)

    def meta_row(label, value):
        return Paragraph(f"<b>{label}:</b> {value}", normal)

    story.append(meta_row("Issued By", "SpacePoint FZC"))
    story.append(meta_row("Date", letter_date))
    story.append(meta_row("Facilitator", instructor.name))
    story.append(meta_row("Reference Agreement", letter.reference or "Facilitator Agreement"))
    story.append(Spacer(1, 6 * mm))

    # ── Subject line ─────────────────────────────────────────────────────
    subject_style = ParagraphStyle("subject", fontSize=10, fontName="Helvetica-Bold", leading=14)
    body_style = ParagraphStyle("body", fontSize=10, fontName="Helvetica", leading=14,
                                textColor=colors.HexColor("#333333"))
    story.append(Paragraph("<b>Subject: Confirmation of Scheduled Workshops and Payment Terms</b>", subject_style))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(
        "According to the Facilitator Agreement referenced above, this letter confirms the scope of work "
        "and agreed compensation for the following workshop sessions delivered by the Facilitator:",
        body_style
    ))
    story.append(Spacer(1, 5 * mm))

    # ── Schedule of Workshops table ───────────────────────────────────────
    story.append(Paragraph("<b>Schedule of Workshops</b>", bold_label))
    story.append(Spacer(1, 2 * mm))

    ws_header = ["Session/Course\nDate", "Workshop\nDescription", "Role", "Location", "Duration", "Compensation\n(AED)"]
    ws_data = [ws_header]
    for s in sessions:
        ws_data.append([
            s.session_date,
            s.workshop_description,
            s.role.value if hasattr(s.role, "value") else str(s.role),
            s.location,
            f"{s.duration_hours:.0f} hours" if s.duration_hours == int(s.duration_hours) else f"{s.duration_hours} hours",
            f"{s.compensation_aed:,.0f}",
        ])

    ws_col_widths = [28 * mm, 40 * mm, 32 * mm, 38 * mm, 18 * mm, 22 * mm]
    ws_table = Table(ws_data, colWidths=ws_col_widths, repeatRows=1)
    ws_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), TABLE_HEADER_GREY),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#aaaaaa")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),   # compensation column right-aligned
        ("ALIGN", (-2, 0), (-2, -1), "CENTER"),  # duration center
    ]))
    story.append(ws_table)
    story.append(Spacer(1, 5 * mm))

    # ── Optional Add-ons table ────────────────────────────────────────────
    if addons:
        story.append(Paragraph("<b>Optional Add-ons</b>", bold_label))
        story.append(Spacer(1, 2 * mm))

        ao_header = ["Description", "Amount (AED)", "Notes"]
        ao_data = [ao_header]
        for a in addons:
            ao_data.append([
                a.description,
                f"{a.amount_aed:,.0f}",
                a.notes or "",
            ])

        ao_col_widths = [65 * mm, 35 * mm, 68 * mm]
        ao_table = Table(ao_data, colWidths=ao_col_widths, repeatRows=1)
        ao_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), TABLE_HEADER_GREY),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#aaaaaa")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        story.append(ao_table)
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(f"<b>Total Optional Add-ons: AED {addons_total:,.0f}</b>", body_style))
        story.append(Spacer(1, 4 * mm))

    # ── Total Compensation ────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    story.append(Spacer(1, 3 * mm))
    total_style = ParagraphStyle("total", fontSize=11, fontName="Helvetica-Bold",
                                 textColor=BRAND_DARK, leading=16)
    story.append(Paragraph(f"<b>Total Compensation Payable: AED {grand_total:,.0f}</b>", total_style))
    story.append(Spacer(1, 2 * mm))

    breakdown_parts = []
    if sessions:
        breakdown_parts.append(f"base compensation of AED {base_total:,.0f}")
    if addons:
        breakdown_parts.append(f"optional add-ons of AED {addons_total:,.0f}")
    if breakdown_parts:
        story.append(Paragraph(f"<i>Total includes {' + '.join(breakdown_parts)}.</i>", body_style))

    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        "Payment will be made by bank transfer within 30 working days after submission of this signed "
        "letter, subject to verification.",
        body_style
    ))
    story.append(Spacer(1, 6 * mm))

    # ── Bank Details ──────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph("<b>Facilitator Bank Details</b>", bold_label))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph("Please confirm or update your current banking details below:", body_style))
    story.append(Spacer(1, 2 * mm))

    if bank:
        bank_items = [
            ("Account Holder Name", bank.account_holder_name or "—"),
            ("Bank Name", bank.bank_name or "—"),
            ("IBAN", bank.iban or "—"),
            ("SWIFT/BIC (if applicable)", bank.swift_bic or "—"),
        ]
        for label, val in bank_items:
            story.append(Paragraph(f"• <b>{label}:</b> {val}", body_style))
    else:
        story.append(Paragraph("• <i>Bank details not yet provided by instructor.</i>", body_style))

    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        "<i>Note: Any changes in bank details must be communicated to SpacePoint in writing.</i>",
        ParagraphStyle("italic_note", fontSize=9, fontName="Helvetica-Oblique",
                       textColor=colors.HexColor("#555555"), leading=13)
    ))
    story.append(Spacer(1, 6 * mm))

    # ── Signatures ────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph("<b>Signatures</b>", bold_label))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        "By signing below, the Parties confirm agreement to the terms of this Payment Letter.",
        body_style
    ))
    story.append(Spacer(1, 4 * mm))

    # Signature table: SpacePoint FZC | Facilitator
    sp_sig_cell = []
    sp_sig_cell.append(Paragraph("<b>For SpacePoint FZC</b>", body_style))
    sp_sig_cell.append(Spacer(1, 1 * mm))
    sp_sig_cell.append(Paragraph(f"Name: {admin_name}", body_style))
    sp_sig_cell.append(Spacer(1, 1 * mm))

    # Admin signature image
    if admin_sig_path and os.path.exists(admin_sig_path):
        try:
            sig_img = RLImage(admin_sig_path, width=30 * mm, height=12 * mm)
            sp_sig_cell.append(sig_img)
        except Exception:
            sp_sig_cell.append(Paragraph("Signature: _______________", body_style))
    else:
        sp_sig_cell.append(Paragraph("Signature: _______________", body_style))

    sp_sig_cell.append(Spacer(1, 1 * mm))
    sp_sig_cell.append(Paragraph(f"Date: {letter_date}", body_style))

    # Instructor signature cell
    inst_sig_cell = []
    inst_sig_cell.append(Paragraph("<b>Facilitator</b>", body_style))
    inst_sig_cell.append(Spacer(1, 1 * mm))
    inst_sig_cell.append(Paragraph(f"Name: {instructor.name}", body_style))
    inst_sig_cell.append(Spacer(1, 1 * mm))

    # Embedded instructor signature if signed
    if letter.instructor_signature_data:
        try:
            sig_bytes = base64.b64decode(letter.instructor_signature_data.split(",")[-1])
            sig_buf = io.BytesIO(sig_bytes)
            inst_sig_img = RLImage(sig_buf, width=30 * mm, height=12 * mm)
            inst_sig_cell.append(inst_sig_img)
        except Exception:
            inst_sig_cell.append(Paragraph("Signature: _______________", body_style))
    else:
        inst_sig_cell.append(Paragraph("Signature: _______________", body_style))

    inst_sig_cell.append(Spacer(1, 1 * mm))
    signed_date = ""
    if letter.signed_at:
        signed_date = letter.signed_at.strftime("%d/%m/%Y")
    inst_sig_cell.append(Paragraph(f"Date: {signed_date or '_______________'}", body_style))

    sig_table = Table([[sp_sig_cell, inst_sig_cell]], colWidths=[85 * mm, 85 * mm])
    sig_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(sig_table)

    # ── Build ─────────────────────────────────────────────────────────────
    try:
        doc.build(story, onFirstPage=draw_full_header, onLaterPages=draw_full_header)
    except Exception as e:
        print(f"[payment_service] PDF generation failed: {e}")
        return None

    # Update DB record
    letter.pdf_path = output_path
    db.add(letter)
    db.commit()
    db.refresh(letter)

    print(f"[payment_service] Generated PDF: {output_path}")
    return output_path


# ── embed instructor signature and regenerate PDF ──────────────────────────
def embed_instructor_signature(letter_id: int, signature_b64: str, db: Session) -> Optional[str]:
    """
    Stores the instructor signature and regenerates the PDF with it embedded.
    Returns the path to the signed PDF or None on failure.
    """
    letter = db.query(PaymentLetter).filter(PaymentLetter.id == letter_id).first()
    if not letter:
        return None

    letter.instructor_signature_data = signature_b64
    letter.signed_at = datetime.utcnow()
    letter.status = PaymentLetterStatus.SIGNED
    db.add(letter)
    db.commit()

    # Regenerate the PDF (now with signature embedded)
    pdf_path = generate_payment_letter_pdf(letter_id, db)

    if pdf_path:
        # Rename as the "signed" version
        signed_path = pdf_path.replace(".pdf", "_SIGNED.pdf")
        import shutil
        shutil.copy2(pdf_path, signed_path)
        letter.signed_pdf_path = signed_path
        db.add(letter)
        db.commit()
        return signed_path

    return None


# ── Excel bulk import parser ───────────────────────────────────────────────
def parse_excel_bulk_import(file_bytes: bytes):
    """
    Parses a 2-sheet xlsx file:
      Sheet 1 "Sessions"  → workshop session rows
      Sheet 2 "Add-ons"   → optional add-on rows (optional sheet)

    Returns a dict:
    {
      "instructors": {
        "email@x.com": {
          "sessions": [ {...}, ... ],
          "addons":   [ {...}, ... ]
        }
      },
      "errors": [ "row N: message", ... ]
    }
    """
    try:
        import openpyxl
    except ImportError:
        return {"instructors": {}, "errors": ["openpyxl not installed. Run: pip install openpyxl"]}

    result = {"instructors": {}, "errors": []}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open file: {e}")
        return result

    SESSIONS_SHEET = "Sessions"
    ADDONS_SHEET = "Add-ons"

    # ── Sheet 1: Sessions ─────────────────────────────────────────────────
    if SESSIONS_SHEET not in wb.sheetnames:
        result["errors"].append(f'Missing required sheet named "{SESSIONS_SHEET}".')
        return result

    ws_sessions = wb[SESSIONS_SHEET]
    rows = list(ws_sessions.iter_rows(values_only=True))
    if not rows:
        result["errors"].append("Sessions sheet is empty.")
        return result

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    REQUIRED_COLS = [
        "instructor email", "session date", "workshop description",
        "role", "location", "duration (hours)", "compensation (aed)"
    ]

    def find_col(name):
        try:
            return headers.index(name.lower())
        except ValueError:
            return None

    col_email = find_col("instructor email")
    col_date = find_col("session date")
    col_desc = find_col("workshop description")
    col_role = find_col("role")
    col_location = find_col("location")
    col_duration = find_col("duration (hours)")
    col_comp = find_col("compensation (aed)")

    missing = [c for c, idx in [
        ("Instructor Email", col_email), ("Session Date", col_date),
        ("Workshop Description", col_desc), ("Role", col_role),
        ("Location", col_location), ("Duration (hours)", col_duration),
        ("Compensation (AED)", col_comp)
    ] if idx is None]
    if missing:
        result["errors"].append(f"Sessions sheet missing columns: {', '.join(missing)}")
        return result

    VALID_ROLES = {
        "lead facilitator": "Lead Facilitator",
        "facilitator": "Facilitator",
        "assistant facilitator": "Assistant Facilitator",
    }

    for row_idx, row in enumerate(rows[1:], start=2):
        email = str(row[col_email]).strip() if row[col_email] else ""
        if not email or email.lower() == "none":
            continue  # skip blank rows

        session_date = str(row[col_date]).strip() if row[col_date] else ""
        workshop_desc = str(row[col_desc]).strip() if row[col_desc] else ""
        role_raw = str(row[col_role]).strip() if row[col_role] else ""
        location = str(row[col_location]).strip() if row[col_location] else ""

        try:
            duration = float(row[col_duration]) if row[col_duration] is not None else 0
        except (ValueError, TypeError):
            result["errors"].append(f"Row {row_idx}: Duration must be a number, got '{row[col_duration]}'")
            duration = 0

        try:
            compensation = float(row[col_comp]) if row[col_comp] is not None else 0
        except (ValueError, TypeError):
            result["errors"].append(f"Row {row_idx}: Compensation must be a number, got '{row[col_comp]}'")
            compensation = 0

        role = VALID_ROLES.get(role_raw.lower())
        if not role:
            result["errors"].append(
                f"Row {row_idx}: Invalid role '{role_raw}'. Must be one of: {', '.join(VALID_ROLES.values())}"
            )
            role = "Facilitator"

        if email not in result["instructors"]:
            result["instructors"][email] = {"sessions": [], "addons": []}

        result["instructors"][email]["sessions"].append({
            "session_date": session_date,
            "workshop_description": workshop_desc,
            "role": role,
            "location": location,
            "duration_hours": duration,
            "compensation_aed": compensation,
        })

    # ── Sheet 2: Add-ons (optional) ───────────────────────────────────────
    if ADDONS_SHEET in wb.sheetnames:
        ws_addons = wb[ADDONS_SHEET]
        ao_rows = list(ws_addons.iter_rows(values_only=True))
        if ao_rows:
            ao_headers = [str(h).strip().lower() if h else "" for h in ao_rows[0]]
            ao_email = None
            ao_desc = None
            ao_amount = None
            ao_notes = None
            try:
                ao_email = ao_headers.index("instructor email")
                ao_desc = ao_headers.index("description")
                ao_amount = ao_headers.index("amount (aed)")
                ao_notes = ao_headers.index("notes") if "notes" in ao_headers else None
            except ValueError as e:
                result["errors"].append(f"Add-ons sheet missing column: {e}")

            if ao_email is not None and ao_desc is not None and ao_amount is not None:
                for row_idx, row in enumerate(ao_rows[1:], start=2):
                    email = str(row[ao_email]).strip() if row[ao_email] else ""
                    if not email or email.lower() == "none":
                        continue
                    desc = str(row[ao_desc]).strip() if row[ao_desc] else ""
                    try:
                        amount = float(row[ao_amount]) if row[ao_amount] is not None else 0
                    except (ValueError, TypeError):
                        result["errors"].append(f"Add-ons row {row_idx}: Amount must be a number.")
                        amount = 0
                    notes = str(row[ao_notes]).strip() if ao_notes is not None and row[ao_notes] else ""

                    if email not in result["instructors"]:
                        result["instructors"][email] = {"sessions": [], "addons": []}
                    result["instructors"][email]["addons"].append({
                        "description": desc,
                        "amount_aed": amount,
                        "notes": notes,
                    })

    return result


# ── generate downloadable Excel template ───────────────────────────────────
def generate_excel_template() -> bytes:
    """Returns the bytes of a blank Excel import template."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Sheet 1 — Sessions
        ws1 = wb.active
        ws1.title = "Sessions"
        sessions_headers = [
            "Instructor Email", "Session Date", "Workshop Description",
            "Role", "Location", "Duration (hours)", "Compensation (AED)"
        ]
        for col, header in enumerate(sessions_headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="653f84")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 22

        # Example row
        example = [
            "instructor@example.com", "30/10/2025", "Quick Flight Workshop",
            "Facilitator", "GEMS Metropole School, Dubai", 4, 500
        ]
        for col, val in enumerate(example, 1):
            ws1.cell(row=2, column=col, value=val)

        # Sheet 2 — Add-ons
        ws2 = wb.create_sheet("Add-ons")
        addons_headers = ["Instructor Email", "Description", "Amount (AED)", "Notes"]
        for col, header in enumerate(addons_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="653f84")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 28

        # Example add-on row
        addon_example = ["instructor@example.com", "Travel allowance", 150, "For session on 30/10"]
        for col, val in enumerate(addon_example, 1):
            ws2.cell(row=2, column=col, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception as e:
        print(f"[payment_service] Failed to generate Excel template: {e}")
        return b""
